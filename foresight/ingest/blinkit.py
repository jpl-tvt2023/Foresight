"""Blinkit adapters: monthly payout zip (13 xlsx files) and daily sales export."""
from __future__ import annotations

import io
import re
import zipfile
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl

from foresight import db

ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
PAREN_RE = re.compile(r"\s*\(.*?\)\s*")


# ---------------------------------------------------------------- helpers

def _norm_state(raw) -> str | None:
    """'Andhra Pradesh (Newly Added)' -> 'Andhra Pradesh'; also used for
    cities ('Rampur (Uttar Pradesh)' -> 'Rampur')."""
    if raw is None:
        return None
    s = PAREN_RE.sub(" ", str(raw)).strip()
    if not s or s == "-":
        return None
    return s


def _parse_date(raw) -> str | None:
    """Accept datetime, '1 May 2026', '2026-05-01', '01-May-2026' -> ISO string."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = str(raw).strip()
    if not s or s == "-":
        return None
    if ISO_DATE_RE.match(s):
        return s[:10]
    for fmt in ("%d %B %Y", "%d %b %Y", "%d-%B-%Y", "%d-%b-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _num(raw, default=0.0) -> float:
    if raw is None:
        return default
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace(",", "").strip()
    if not s or s in ("-", "NA", "N/A"):
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _sheet_rows(ws, anchor: str):
    """Yield dict rows keyed by header, finding the header row by an anchor column name."""
    header, header_idx = None, None
    for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if anchor in cells:
            header, header_idx = cells, i
            break
    if header is None:
        return
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        yield dict(zip(header, row))


def _load_ws(zf: zipfile.ZipFile, member: str, sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(member)), read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None, None
    return wb, wb[sheet]


def _sheet_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """Map sheet name -> zip member. Blinkit renamed the files across export
    versions (e.g. 'Storage Charges.xlsx' vs 'Ageing Inventory.xlsx') but the
    sheet names are stable, so discovery goes by sheet.

    'Payout Breakup' appears as a copy in every file; prefer the dedicated
    single-sheet member for it (any copy would do).
    """
    smap: dict[str, str] = {}
    for m in zf.namelist():
        if not m.lower().endswith(".xlsx"):
            continue
        try:
            wb = openpyxl.load_workbook(io.BytesIO(zf.read(m)), read_only=True)
            sheets = list(wb.sheetnames)
            wb.close()
        except Exception:
            continue
        for s in sheets:
            if s == "Payout Breakup" and len(sheets) > 1:
                smap.setdefault(s, m)      # keep dedicated file if already seen
            elif s == "Payout Breakup":
                smap[s] = m                # the dedicated single-sheet file wins
            else:
                smap.setdefault(s, m)
    return smap


def _open_sheet(zf: zipfile.ZipFile, smap: dict[str, str], sheet: str):
    member = smap.get(sheet)
    if not member:
        return None, None
    return _load_ws(zf, member, sheet)


# ---------------------------------------------------------------- monthly zip

def ingest_monthly_zip(conn, zip_path: str | Path, cycle_label: str | None = None) -> dict:
    """Ingest one Blinkit monthly payout zip. Idempotent per cycle_label (skips if seen)."""
    zip_path = Path(zip_path)
    if cycle_label is None:
        m = re.search(r"(\d{4}-\d{2})", zip_path.name)
        cycle_label = m.group(1) if m else zip_path.stem

    pid = db.get_platform_id(conn)
    existing = conn.execute(
        "SELECT id FROM payout_cycles WHERE platform_id=? AND cycle_label=?",
        (pid, cycle_label)).fetchone()
    if existing:
        return {"cycle": cycle_label, "skipped": "already ingested"}

    stats: dict = {"cycle": cycle_label}
    with zipfile.ZipFile(zip_path) as zf:
        smap = _sheet_map(zf)
        if "Forward Orders" not in smap:
            raise ValueError(
                f"{zip_path.name}: no 'Forward Orders' sheet in any file — "
                "not a Blinkit payout zip (or a new export format); nothing ingested")

        period_start = period_end = None
        wb, ws = _open_sheet(zf, smap, "Note")
        if ws:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None]
                for i, v in enumerate(vals):
                    if v == "Payout Cycle" and i + 1 < len(vals):
                        parts = vals[i + 1].split(" to ")
                        if len(parts) == 2:
                            period_start = _parse_date(parts[0])
                            period_end = _parse_date(parts[1])
            wb.close()

        cur = conn.execute(
            "INSERT INTO payout_cycles(platform_id, cycle_label, period_start, period_end, source_file) "
            "VALUES (?,?,?,?,?)",
            (pid, cycle_label, period_start, period_end, zip_path.name))
        cycle_id = cur.lastrowid

        loc_cache: dict[str, int] = {}

        def loc_id(state: str) -> int:
            if state not in loc_cache:
                loc_cache[state] = db.upsert_location(conn, pid, state)
            return loc_cache[state]

        warehouse_state: dict[str, str] = {}

        stats |= _ingest_forward_orders(conn, zf, smap, pid)
        stats |= _ingest_returns(conn, zf, smap, pid)
        stats |= _ingest_ageing(conn, zf, smap, pid, cycle_id, loc_id)
        stats |= _ingest_upfront_grn(conn, zf, smap, pid, cycle_id, loc_id, warehouse_state)
        stats |= _ingest_recall(conn, zf, smap, pid, cycle_id, loc_id, warehouse_state)
        stats |= _ingest_lost_damaged(conn, zf, smap, pid, cycle_id, loc_id)
        stats |= _ingest_simple_charges(conn, zf, smap, cycle_id)
        stats |= _ingest_payout_breakup(conn, zf, smap, cycle_id)

    conn.commit()
    return stats


def _ingest_forward_orders(conn, zf, smap, pid) -> dict:
    wb, ws = _open_sheet(zf, smap, "Forward Orders")
    if ws is None:
        return {"forward_orders": "sheet missing"}

    # Decision grain is SKU x city (spec): Customer City == the dark-store city
    # in q-commerce, and Supply State == Customer State on 100% of verified rows.
    agg = defaultdict(lambda: [0, 0.0])           # (item, state, city, date) -> [units, gross]
    margin = defaultdict(lambda: [0.0, 0])        # item -> [payout_sum, units]
    item_meta: dict[str, dict] = {}
    rows = 0
    for r in _sheet_rows(ws, "Item ID"):
        d = _parse_date(r.get("Order Date"))
        item = r.get("Item ID")
        state = _norm_state(r.get("Supply State"))
        if not (d and item and state):
            continue
        city = _norm_state(r.get("Customer City"))
        units = int(_num(r.get("Quantity")))
        gross = _num(r.get("Total Gross Bill Amount"))
        agg[(str(item), state, city, d)][0] += units
        agg[(str(item), state, city, d)][1] += gross
        payout = _num(r.get("Item Level Payout"))
        if payout and units:
            margin[str(item)][0] += payout
            margin[str(item)][1] += units
        if str(item) not in item_meta:
            item_meta[str(item)] = dict(
                name=r.get("Product Name"), variant=r.get("Variant Description"),
                l0_category=r.get("L0 Category"), l1_category=r.get("L1 Category"),
                l2_category=r.get("L2 Category"), hsn_code=str(r.get("HSN Code") or "") or None)
        rows += 1
    wb.close()

    item_ids: dict[str, int] = {}
    for pit, meta in item_meta.items():
        m = margin.get(pit)
        est = round(m[0] / m[1], 2) if m and m[1] else None
        item_ids[pit] = db.upsert_item(conn, pid, pit, unit_margin_estimate=est, **meta)

    loc_ids: dict[tuple, int] = {}
    for (pit, state, city, d), (units, gross) in agg.items():
        if (state, city) not in loc_ids:
            loc_ids[(state, city)] = db.upsert_location(conn, pid, state, city)
        iid = item_ids.get(pit) or db.upsert_item(conn, pid, pit)
        item_ids[pit] = iid
        lid = loc_ids[(state, city)]
        conn.execute(
            "INSERT INTO sales_daily(platform_id, item_id, location_id, sale_date, units, gross_value) "
            "VALUES (?,?,?,?,?,?) "
            "ON CONFLICT(platform_id, item_id, location_id, sale_date) "
            "DO UPDATE SET units = units + excluded.units, gross_value = gross_value + excluded.gross_value",
            (pid, iid, lid, d, units, gross))
        conn.execute(
            "INSERT INTO stock_ledger(platform_id, item_id, location_id, event_date, event_type, units_delta, note) "
            "VALUES (?,?,?,?, 'sale', ?, ?)",
            (pid, iid, lid, d, -units, "forward orders"))
    return {"forward_order_rows": rows, "sales_cells": len(agg)}


def _ingest_returns(conn, zf, smap, pid) -> dict:
    wb, ws = _open_sheet(zf, smap, "Cancelled or Returned Orders")
    if ws is None:
        return {"returns": "sheet missing"}
    agg = defaultdict(int)
    for r in _sheet_rows(ws, "Item ID"):
        d = _parse_date(r.get("Return Order Date"))
        item = r.get("Item ID")
        state = _norm_state(r.get("Supply State"))
        if not (d and item and state):
            continue
        city = _norm_state(r.get("Customer City"))
        agg[(str(item), state, city, d)] += int(_num(r.get("Quantity")))
    wb.close()
    for (pit, state, city, d), units in agg.items():
        iid = db.upsert_item(conn, pid, pit)
        lid = db.upsert_location(conn, pid, state, city)
        conn.execute(
            "INSERT INTO sales_daily(platform_id, item_id, location_id, sale_date, units, gross_value, returns_units) "
            "VALUES (?,?,?,?,0,0,?) "
            "ON CONFLICT(platform_id, item_id, location_id, sale_date) "
            "DO UPDATE SET returns_units = returns_units + excluded.returns_units",
            (pid, iid, lid, d, units))
        conn.execute(
            "INSERT INTO stock_ledger(platform_id, item_id, location_id, event_date, event_type, units_delta, note) "
            "VALUES (?,?,?,?, 'return', ?, 'returned units re-enter stock')",
            (pid, iid, lid, d, units))
    return {"return_cells": len(agg)}


def _ingest_ageing(conn, zf, smap, pid, cycle_id, loc_id) -> dict:
    """Daily Ageing (wide: one row per state x item x slab, date columns)."""
    wb, ws = _open_sheet(zf, smap, "Daily Ageing")
    ageing_rows = 0
    if ws is not None:
        header = None
        for row in ws.iter_rows(max_row=8, values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if "Item ID" in cells and "State" in cells:
                header = cells
                break
        if header:
            date_cols = [(i, c) for i, c in enumerate(header) if ISO_DATE_RE.match(c)]
            idx = {name: header.index(name) for name in
                   ("State", "Item ID", "Item Name", "Per day charge (Rs)", "Ageing Slab", "Total Charge (Rs)")}
            started = False
            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                if not started:
                    cells = [str(c).strip() if c is not None else "" for c in vals]
                    if cells == header:
                        started = True
                    continue
                state = _norm_state(vals[idx["State"]])
                item = vals[idx["Item ID"]]
                if not (state and item):
                    continue
                iid = db.upsert_item(conn, pid, str(item), name=vals[idx["Item Name"]])
                lid = loc_id(state)
                per_day = _num(vals[idx["Per day charge (Rs)"]])
                slab = str(vals[idx["Ageing Slab"]] or "")
                for ci, cdate in date_cols:
                    units = vals[ci]
                    if units is None:
                        continue
                    u = int(_num(units))
                    conn.execute(
                        "INSERT OR REPLACE INTO storage_ageing"
                        "(cycle_id, item_id, location_id, ageing_date, units, per_day_charge, ageing_slab, charge_amount) "
                        "VALUES (?,?,?,?,?,?,?,?)",
                        (cycle_id, iid, lid, cdate, u, per_day, slab, round(u * per_day, 2)))
                total = _num(vals[idx["Total Charge (Rs)"]])
                if total:
                    conn.execute(
                        "INSERT INTO charges(cycle_id, charge_type, item_id, location_id, amount, note) "
                        "VALUES (?,?,?,?,?,?)",
                        (cycle_id, "storage", iid, lid, total, slab))
                ageing_rows += 1
    if wb is not None:
        wb.close()
    return {"ageing_rows": ageing_rows}


def _ingest_upfront_grn(conn, zf, smap, pid, cycle_id, loc_id, warehouse_state) -> dict:
    """Upfront Storage Charges = GRN receipts (also builds the warehouse->state map)."""
    wb, ws = _open_sheet(zf, smap, "Upfront Storage Charges")
    grn_rows = 0
    if ws is not None:
        for r in _sheet_rows(ws, "STO Number"):
            state = _norm_state(r.get("State"))
            item = r.get("Item ID")
            d = _parse_date(r.get("GRN Date"))
            if not (state and item and d):
                continue
            wh = str(r.get("Warehouse") or "").strip()
            if wh:
                warehouse_state[wh] = state
            iid = db.upsert_item(conn, pid, str(item), name=r.get("Item Name"))
            lid = loc_id(state)
            units = int(_num(r.get("Qty")))
            sto = str(r.get("STO Number") or "")
            conn.execute(
                "INSERT OR IGNORE INTO replenishments"
                "(platform_id, item_id, location_id, expected_live_date, units, status, sto_number) "
                "VALUES (?,?,?,?,?,'live',?)",
                (pid, iid, lid, d, units, sto))
            conn.execute(
                "INSERT INTO stock_ledger(platform_id, item_id, location_id, event_date, event_type, units_delta, note) "
                "VALUES (?,?,?,?, 'receipt', ?, ?)",
                (pid, iid, lid, d, units, f"GRN {sto} @ {wh}"))
            amt = _num(r.get("Total Charge (Rs)"))
            if amt:
                conn.execute(
                    "INSERT INTO charges(cycle_id, charge_type, item_id, location_id, charge_date, units, amount, note) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (cycle_id, "upfront_storage", iid, lid, d, units, amt, wh))
            grn_rows += 1
    if wb is not None:
        wb.close()
    return {"grn_receipts": grn_rows}


def _state_from_warehouse(wh: str, warehouse_state: dict) -> str | None:
    if wh in warehouse_state:
        return warehouse_state[wh]
    city = wh.split(" - ")[0].strip()
    city = re.sub(r"\s+[A-Z]\d+$", "", city)  # 'Bengaluru B3' -> 'Bengaluru'
    for known, state in warehouse_state.items():
        if known.startswith(city):
            return state
    return None


def _ingest_recall(conn, zf, smap, pid, cycle_id, loc_id, warehouse_state) -> dict:
    wb, ws = _open_sheet(zf, smap, "Recall Charge")
    if ws is None:
        return {}
    n = 0
    for r in _sheet_rows(ws, "Recall Batch ID"):
        item = r.get("Item ID")
        if not item:
            continue
        d = _parse_date(r.get("Recall Dispatch Date")) or _parse_date(r.get("Recall Requested Date"))
        wh = str(r.get("Warehouse Name") or "")
        state = _state_from_warehouse(wh, warehouse_state)
        iid = db.upsert_item(conn, pid, str(item), name=r.get("Item Name"))
        units = int(_num(r.get("Qty")))
        amt = _num(r.get("Total Charge (Rs)"))
        lid = loc_id(state) if state else None
        conn.execute(
            "INSERT INTO charges(cycle_id, charge_type, item_id, location_id, charge_date, units, amount, note) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (cycle_id, "recall", iid, lid, d, units, amt, wh))
        if state and d:
            conn.execute(
                "INSERT INTO stock_ledger(platform_id, item_id, location_id, event_date, event_type, units_delta, note) "
                "VALUES (?,?,?,?, 'recall', ?, ?)",
                (pid, iid, lid, d, -units, f"recall from {wh}"))
        n += 1
    wb.close()
    return {"recall_rows": n}


def _ingest_lost_damaged(conn, zf, smap, pid, cycle_id, loc_id) -> dict:
    wb, ws = _open_sheet(zf, smap, "Lost Damaged Inventory")
    if ws is None:
        return {}
    n = 0
    for r in _sheet_rows(ws, "Item ID"):
        item = r.get("Item ID")
        state = _norm_state(r.get("State Name"))
        if not (item and state):
            continue
        iid = db.upsert_item(conn, pid, str(item), name=r.get("Product Name"))
        lid = loc_id(state)
        units = int(_num(r.get("Quantity")))
        amt = _num(r.get("Total Compensation Amount"))
        conn.execute(
            "INSERT INTO charges(cycle_id, charge_type, item_id, location_id, units, amount, note) "
            "VALUES (?,?,?,?,?,?, 'compensation credited')",
            (cycle_id, "lost_damaged", iid, lid, units, amt))
        conn.execute(
            "INSERT INTO stock_ledger(platform_id, item_id, location_id, event_date, event_type, units_delta, note) "
            "VALUES (?,?,?, date('now'), 'adjustment', ?, 'lost/damaged inventory')",
            (pid, iid, lid, -units))
        n += 1
    wb.close()
    return {"lost_damaged_rows": n}


def _ingest_simple_charges(conn, zf, smap, cycle_id) -> dict:
    """Flat-total charge sheets for the payout analytics layer."""
    specs = [
        ("Return Charges", "return", "Return Charge", "Return Charge GST"),
        ("Courier Charge", "courier", "Amount", None),
        ("Credit Note-Debit Note", "cn_dn", "Total amount", None),
    ]
    out = {}
    for sheet, ctype, amount_col, gst_col in specs:
        wb, ws = _open_sheet(zf, smap, sheet)
        if ws is None:
            continue
        n = 0
        for r in _sheet_rows(ws, "S.No."):
            amt = _num(r.get(amount_col))
            if not amt:
                continue
            gst = _num(r.get(gst_col)) if gst_col else 0.0
            d = _parse_date(r.get("Delivery Date") or r.get("CN/DN issue date")
                            or r.get("Return Invoice Date"))
            conn.execute(
                "INSERT INTO charges(cycle_id, charge_type, charge_date, amount, gst_amount, note) "
                "VALUES (?,?,?,?,?,?)",
                (cycle_id, ctype, d, amt, gst,
                 str(r.get("Return Reason") or r.get("Description") or r.get("Courier Name") or "")))
            n += 1
        wb.close()
        out[f"{ctype}_rows"] = n
    return out


def _ingest_payout_breakup(conn, zf, smap, cycle_id) -> dict:
    wb, ws = _open_sheet(zf, smap, "Payout Breakup")
    if ws is None:
        return {}
    n = 0
    for r in _sheet_rows(ws, "Particular"):
        p = r.get("Particular")
        if not p:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO payout_summary(cycle_id, particular, delivered_amount, cancelled_returned_amount, total_amount) "
            "VALUES (?,?,?,?,?)",
            (cycle_id, str(p), _num(r.get("Delivered Orders"), None),
             _num(r.get("Cancelled/Returned"), None), _num(r.get("Total"), None)))
        n += 1
    wb.close()
    return {"payout_particulars": n}


# ---------------------------------------------------------------- daily sales export

DAILY_COL_ALIASES = {
    "date": ["Order Date", "Date", "order_date", "sale_date"],
    "item": ["Item ID", "item_id", "SKU", "Product ID", "product_id"],
    "state": ["Supply State", "State", "Customer State", "state"],
    "city": ["Customer City", "City", "city", "Supply City"],
    "units": ["Quantity", "Units", "Qty", "units", "quantity"],
    "gross": ["Total Gross Bill Amount", "GMV", "Gross Value", "Selling Price (Rs)", "gross_value"],
    "name": ["Product Name", "Item Name", "name"],
}


def ingest_daily_sales(conn, path: str | Path) -> dict:
    """Flexible adapter for the Blinkit seller-panel daily sales export (csv/xlsx).

    Column names are matched against DAILY_COL_ALIASES; exact schema locks in once a
    sample export is provided (spec §10.2).
    """
    import pandas as pd
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xls"):
        frame = pd.read_excel(path)
        if not any(a in frame.columns for a in DAILY_COL_ALIASES["item"]):
            # header may not be the first row (Blinkit banner rows) — rescan
            raw = pd.read_excel(path, header=None)
            for i in range(min(10, len(raw))):
                if any(str(v).strip() in DAILY_COL_ALIASES["item"] for v in raw.iloc[i]):
                    frame = pd.read_excel(path, header=i)
                    break
    else:
        frame = pd.read_csv(path)

    cols = {}
    for key, aliases in DAILY_COL_ALIASES.items():
        for a in aliases:
            if a in frame.columns:
                cols[key] = a
                break
    missing = {"date", "item", "state", "units"} - set(cols)
    if missing:
        raise ValueError(f"daily sales export missing columns: {missing}; found {list(frame.columns)}")

    pid = db.get_platform_id(conn)
    agg = defaultdict(lambda: [0, 0.0])
    names = {}
    for _, r in frame.iterrows():
        d = _parse_date(r[cols["date"]])
        item = r[cols["item"]]
        state = _norm_state(r[cols["state"]])
        if not (d and state) or item is None:
            continue
        city = _norm_state(r[cols["city"]]) if "city" in cols else None
        item = str(item).split(".")[0]
        agg[(item, state, city, d)][0] += int(_num(r[cols["units"]]))
        if "gross" in cols:
            agg[(item, state, city, d)][1] += _num(r[cols["gross"]])
        if "name" in cols and item not in names:
            names[item] = r[cols["name"]]

    for (pit, state, city, d), (units, gross) in agg.items():
        iid = db.upsert_item(conn, pid, pit, name=names.get(pit))
        lid = db.upsert_location(conn, pid, state, city)
        conn.execute(
            "INSERT INTO sales_daily(platform_id, item_id, location_id, sale_date, units, gross_value) "
            "VALUES (?,?,?,?,?,?) "
            "ON CONFLICT(platform_id, item_id, location_id, sale_date) "
            "DO UPDATE SET units = excluded.units, gross_value = excluded.gross_value",
            (pid, iid, lid, d, units, gross))
        conn.execute(
            "INSERT INTO stock_ledger(platform_id, item_id, location_id, event_date, event_type, units_delta, note) "
            "VALUES (?,?,?,?, 'sale', ?, 'daily sales export')",
            (pid, iid, lid, d, -units))
    conn.commit()
    return {"daily_sales_cells": len(agg)}
